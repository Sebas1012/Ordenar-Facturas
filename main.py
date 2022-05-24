from openpyxl import load_workbook
import json
import shutil

def move_invoices(book_path, sheet_name, json_path, invoices_path):
  # Excel file load
  book = load_workbook(str(book_path))
  
  if sheet_name == 'None':
    sheet = book.active
  else:
    sheet = book[str(sheet_name)]

  # JSON path file load
  with open(json_path) as data:
    data = json.load(data)

  for row in sheet.iter_rows(min_row=2, values_only=True):
    try:
      shutil.move(invoices_path+str(row[0])+data['Extension'], data[row[1]])
      print(str(row[0])+data['Extension']+' movido con exito.')
    except:
      print(str(row[0])+data['Extension']+' no pudo moverse')

book_p      = input("Ingrese la ruta del archivo Excel: ")
sheet_n     = input("Ingrese el nombre de la hoja: ")
json_p      = input("Ingrese la ruta del archivo JSON: ")
invoices_p  = input("Ingrese la ruta de los archivos a ordenar: ")

move_invoices(book_p, sheet_n, json_p, invoices_p)


